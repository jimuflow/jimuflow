# This software is dual-licensed under the GNU General Public License (GPL) 
# and a commercial license.
#
# You may use this software under the terms of the GNU GPL v3 (or, at your option,
# any later version) as published by the Free Software Foundation. See 
# <https://www.gnu.org/licenses/> for details.
#
# If you require a proprietary/commercial license for this software, please 
# contact us at jimuflow@gmail.com for more information.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.
#
# Copyright (C) 2024-2025  Weng Jing

import csv
import os
import tempfile

import openpyxl
import pytest
import xlrd

from jimuflow.components.table import ExportTableComponent
from jimuflow.datatypes import Table
from jimuflow.runtime.execution_engine import ControlFlow
from jimuflow.runtime.expression import escape_string
from tests.utils import create_component


@pytest.mark.asyncio
async def test_export_xlsx():
    with tempfile.TemporaryDirectory() as temp_dir:
        table = Table(['ID', "名称"])
        table.rows = [[1, 'a'], [2, 'b'], [3, 'c']]
        component = create_component(ExportTableComponent)
        await component.process.update_variable('t', table)
        component.node.inputs = {
            "table": 't',
            "saveFolder": escape_string(str(temp_dir)),
            "fileFormat": 'xlsx',
            "sheetName": '"测试"',
            "fileNamingType": 'random',
            "exportHeader": True,
        }
        component.node.outputs = {
            "filePath": 'r'
        }
        assert (await component.execute()) == ControlFlow.NEXT
        xlsx_file = component.process.get_variable('r')
        assert xlsx_file.endswith('.xlsx')
        workbook = openpyxl.load_workbook(xlsx_file)
        sheet = workbook['测试']
        rows = []
        for row in sheet.iter_rows(min_row=sheet.min_row, max_row=sheet.max_row, min_col=sheet.min_column,
                                   max_col=sheet.max_column):
            rows.append([cell.value for cell in row])
        assert rows == [['ID', '名称'], [1, 'a'], [2, 'b'], [3, 'c']]


@pytest.mark.asyncio
async def test_export_xls():
    with tempfile.TemporaryDirectory() as temp_dir:
        table = Table(['ID', "名称"])
        table.rows = [[1, 'a'], [2, 'b'], [3, 'c']]
        component = create_component(ExportTableComponent)
        await component.process.update_variable('t', table)
        component.node.inputs = {
            "table": 't',
            "saveFolder": escape_string(str(temp_dir)),
            "fileFormat": 'xls',
            "sheetName": '"测试"',
            "fileNamingType": 'random',
            "exportHeader": True,
        }
        component.node.outputs = {
            "filePath": 'r'
        }
        assert (await component.execute()) == ControlFlow.NEXT
        xls_file = component.process.get_variable('r')
        assert xls_file.endswith('.xls')
        workbook = xlrd.open_workbook(xls_file)
        sheet = workbook.sheet_by_name('测试')
        rows = []
        for row in range(0, sheet.nrows):
            rows.append([sheet.cell_value(row, i) for i in range(sheet.ncols)])
        assert rows == [['ID', '名称'], [1, 'a'], [2, 'b'], [3, 'c']]


@pytest.mark.asyncio
async def test_export_csv():
    with tempfile.TemporaryDirectory() as temp_dir:
        table = Table(['ID', "名称"])
        table.rows = [[1, 'a'], [2, 'b'], [3, 'c']]
        component = create_component(ExportTableComponent)
        await component.process.update_variable('t', table)
        component.node.inputs = {
            "table": 't',
            "saveFolder": escape_string(str(temp_dir)),
            "fileFormat": 'csv',
            "fileEncoding": 'utf-8',
            "useCustomDelimiter": False,
            "fileNamingType": 'random',
            "exportHeader": True,
        }
        component.node.outputs = {
            "filePath": 'r'
        }
        assert (await component.execute()) == ControlFlow.NEXT
        csv_file = component.process.get_variable('r')
        assert csv_file.endswith('.csv')
        with open(csv_file, mode='r', newline='', encoding='utf-8') as file:
            reader = csv.reader(file, delimiter=',')
            rows = []
            for row in reader:
                rows.append(row)
            assert rows == [['ID', '名称'], ['1', 'a'], ['2', 'b'], ['3', 'c']]


@pytest.mark.asyncio
async def test_export_tsv_with_specified_name():
    with tempfile.TemporaryDirectory() as temp_dir:
        table = Table(['ID', "名称"])
        table.rows = [[1, 'a'], [2, 'b'], [3, 'c']]
        component = create_component(ExportTableComponent)
        await component.process.update_variable('t', table)
        component.node.inputs = {
            "table": 't',
            "saveFolder": escape_string(str(temp_dir)),
            "fileFormat": 'csv',
            "fileEncoding": 'utf-8',
            "useCustomDelimiter": True,
            "customDelimiter": escape_string('\t'),
            "fileNamingType": 'custom',
            "customFilename": '"test"',
            "exportHeader": False,
        }
        component.node.outputs = {
            "filePath": 'r'
        }
        assert (await component.execute()) == ControlFlow.NEXT
        csv_file = component.process.get_variable('r')
        assert os.path.basename(csv_file) == 'test.csv'
        with open(csv_file, mode='r', newline='', encoding='utf-8') as file:
            reader = csv.reader(file, delimiter='\t')
            rows = []
            for row in reader:
                rows.append(row)
            assert rows == [['1', 'a'], ['2', 'b'], ['3', 'c']]
